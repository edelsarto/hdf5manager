from concurrent.futures import ProcessPoolExecutor
import openpyxl
from openpyxl import load_workbook
from time import perf_counter

import pandas as pd
import pandas
import numpy as np
from joblib import Parallel, delayed
import time

import numbers

import math

global wb

# filename_with_direc = "excel/Prova_BL-34.xlsx"


def loadOld(filename_with_direc):

    global wb
    ## read the excel file with name
    wb = openpyxl.load_workbook(filename_with_direc, read_only=True, data_only=True)#, data_only=True


def SaveParameters(filename_with_direc):

    loadOld(filename_with_direc)

    sheet = wb[wb.sheetnames[0]]

    # lettura "Parameter"
    parameterArray = []
    # parameterArrayFiltered = []

    for j in range(sheet.max_row):
        # print(j)
        if sheet.cell(row=j + 1, column=1).value == "Parameter":

            print("--- PARAM ROW FOUND --- ", j + 1)

            i = 0
            for param in sheet.iter_rows(
                min_row=j + 1, max_row=j + 1, min_col=i + 2, max_col=2000
            ):

                # print(" - ", i, " - ")
                # print(param)
                for p in param:

                    if p.value != None:  # sheet.cell(row=j+1, column=i+1).value
                        # print(p.value)
                        parameterArray.append(p.value)
                        # print("#########################")
                        # print(i, ") ", p.value)
                        # print("#########################")

                    i += 1
            break
            # break
    # print("################# LISTA PARAMETRI ##################")
    # print(parameterArray)
    print("######################### END SAVE PARAM #########################")
    return parameterArray


def SaveSensorID(filename_with_direc):

    sensorIdList = []

    #loadOld(filename_with_direc)

    sheet = wb[wb.sheetnames[0]]

    for j in range(sheet.max_row):
        # print(j)
        if sheet.cell(row=j + 1, column=1).value == "Sensor ID":

            print("--- SENSOR ID ROW FOUND --- ", j + 1)

            i = 0

            for sensid in sheet.iter_rows(
                min_row=j + 1, max_row=j + 1, min_col=i + 2, max_col=2000
            ):

                # print(" - ", i, " - ")

                for sid in sensid:

                    if (
                        sid.value != None and sid.value != "TIME"
                    ):  # sheet.cell(row=j+1, column=i+1).value

                        sensorIdList.append(sid.value)
                        # print("#########################")
                        # print(i, ") ", sid.value)
                        # print("#########################")

                    i += 1
            break

    print("######################### END SAVE SENSOR ID #########################")

    return sensorIdList


def SaveCategoryList(filename_with_direc):

    #loadOld(filename_with_direc)
    # print(wb.sheetnames)
    sheet = wb[wb.sheetnames[0]]  # SOSTITUIRE ACNHE GLI ALTRI

    categoryList = []

    for j in range(sheet.max_row):

        if sheet.cell(row=j + 1, column=1).value == None:
            break

        categoryList.append(sheet.cell(row=j + 1, column=1).value)
        # print(sheet.cell(row=j+1, column=1).value)

    # print(j)
    return categoryList


def SaveValuesList(filename_with_direc, sensorid):

    start = time.time()

    #loadOld(filename_with_direc)

    sheet = wb[wb.sheetnames[0]]

    array = []

    for i in range(sheet.max_column):

        if i > 1 and sheet.cell(row=1, column=i + 1).value == None:
            break

        if sensorid == sheet.cell(row=1, column=i + 1).value:

            print("SAVING TIME")

            print(i, ") SENSOR ID: " + sheet.cell(row=1, column=i + 1).value)

            for row in sheet.iter_rows(
                min_row=1, max_row=2000, min_col=i + 1, max_col=i + 1
                ):

                for cell in row:

                    #print("##############", i ,") VALUE GET ", cell.value ," ###############")
                    array.append(cell.value)

    end = time.time()
    # print(array)

    print("################### END VALUES SAVE ##################")
    return array


def SaveTimeList(filename_with_direc, sensorid):

    #loadOld(filename_with_direc)

    sheet = wb[wb.sheetnames[0]]

    array = []

    for i in range(sheet.max_column):

        if sheet.cell(row=1, column=i + 1).value == sensorid:

            for column in sheet.iter_rows(
                min_row=1, max_row=2000, min_col=i, max_col=i
            ):
                for cell in column:
                    if (
                        isinstance(cell.value, numbers.Number) == True
                        and cell.value != None
                    ):

                        array.append(cell.value)
                        # print("##############", i ,") TIME GET ", cell.value ," ###############")

            break

    print("################### END TIME SAVE ##################")
    return array

def SaveAndCalculateUncertainty(filename_with_direc, error):
    
    loadOld(filename_with_direc)
    
    sheet = sheet = wb[wb.sheetnames[1]]
    
    a = 0
    b = 0
    c = 0
    d = 0
    
    for i in range(sheet.max_row):
        
        if sheet.cell(row=i+1, column=1).value == error:
             
            a = sheet.cell(row=i+1, column=2).value
            b = sheet.cell(row=i+1, column=3).value
            c = sheet.cell(row=i+1, column=4).value
            d = sheet.cell(row=i+1, column=5).value
            
            break
    
    if(a==0
       and b==0
       and c==0
       and d==0):
        
        return 0
    
    uncertainty = math.sqrt(pow(a,2) + pow(b,2) + pow(c,2) + pow(d,2))
    
    print(uncertainty)
    
    return uncertainty
            
################################## MAIN ###################################################
SaveAndCalculateUncertainty("excel/Prova_BL-34.xlsx","01WH")
# SaveValuesList("excel/Prova_BL-34.xlsx","PA11")
# SaveSensorID("excel/Prova_BL-34.xlsx")
# SaveTimeList("excel/Prova_BL-34.xlsx", "PA11")
# SaveCategoryList("excel/Prova_BL-34.xlsx")
# SaveParameters("excel/Prova_BL-34.xlsx")
# TestIter("excel/Prova_BL-34.xlsx", "PA11")#funzione giusta
# print(SaveSensorID("excel/Prova_BL-34.xlsx"))
